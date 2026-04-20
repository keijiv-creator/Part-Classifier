import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import os
import sys
import time
import json
import re
import statistics
import requests
from collections import defaultdict, Counter
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(os.path.dirname(SCRIPT_DIR))

import argparse

OUTPUT_DIR = os.path.join(ROOT_DIR, 'output')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'Parts_Analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

API_TOKEN = os.environ.get('PIPEDRIVE_API_KEY', '')
BASE_URL = 'https://api.pipedrive.com/v1'

PD_FIELDS = {
    'deal_type':         'ed1ea84ddf0aeccf6d0b3d927026105878826510',
    'mfg_type':          '47d2446f3ab01948599ee675f453635f85a20c9d',
    'industry':          '6aea8f84232712fdf542c789b7d589c271da21fb',
    'customer_part':     'bf83e3810120851228084c7f58f6fa40437c8915',
    'part':              '6609b1b875bcedc0740cbe3aecfe48ef0b17c64f',
    'p1_time':           '4a8eeb233bb88878360ec1d17676db77be8b3b88',
    'p2_time':           'eeb93bef8e9e1d8b8f5d879cd9193f2f9545b693',
    'p3_time':           '504c42fe1681d50c19a37c0204c28b0ade57a51a',
    'p4_time':           'f7c63e1bd65fbc7e80b9001c5f0c1fd8942f80ca',
    'p5_time':           '7fb07387a5fa68572587e7204c546e2f2fca336b',
    'quote_number':      '7253c923d10c16cab1e171ffd0efbf8aee3c1858',
    'po_number':         '05371c14399cb6114fa964998cfae59344f56139',
    'platform_company':  '25e61f29f3d977465c2104addd1a3a6e15107638',
}

OPTION_LABELS = {
    'label': {
        '15': 'New Part', '16': 'New Customer', '23': 'Synergy',
        '24': 'Repeat', '26': 'Dormant', '171': 'New Rev',
        '173': 'OSS - Finish', '174': 'OSS-Material',
        '175': 'OSS - Hardware', '176': 'Tool Check',
        15: 'New Part', 16: 'New Customer', 23: 'Synergy',
        24: 'Repeat', 26: 'Dormant', 171: 'New Rev',
        173: 'OSS - Finish', 174: 'OSS-Material',
        175: 'OSS - Hardware', 176: 'Tool Check',
    },
    PD_FIELDS['industry']: {
        '27': 'A&D', '28': 'Automotive', '31': 'Medical',
        '34': 'Space', '35': 'Industrial', '36': 'EV',
        27: 'A&D', 28: 'Automotive', 31: 'Medical',
        34: 'Space', 35: 'Industrial', 36: 'EV',
    },
    PD_FIELDS['deal_type']: {
        '37': 'Bid', '38': 'Demand',
        37: 'Bid', 38: 'Demand',
    },
    PD_FIELDS['mfg_type']: {
        '41': 'Deep Draw', '42': 'Progressive', '43': 'Secondary',
        '44': 'Swiss', '45': 'Machining', '50': 'Tooling',
        41: 'Deep Draw', 42: 'Progressive', 43: 'Secondary',
        44: 'Swiss', 45: 'Machining', 50: 'Tooling',
    },
    PD_FIELDS['platform_company']: {
        '53': 'Coining', '54': 'GEM', '91': 'National',
        '100': 'MSK', '101': 'Ditron', '140': 'Hudson',
        53: 'Coining', 54: 'GEM', 91: 'National',
        100: 'MSK', 101: 'Ditron', 140: 'Hudson',
    },
}

PHASE_IDS = {
    20: 'Phase 1', 21: 'Phase 2', 22: 'Phase 3', 23: 'Phase 4', 24: 'Phase 5',
    11: 'Inquiry', 17: 'Ready For Pipeline', 26: 'No-Quote',
    35: 'RFQ Received', 36: 'In Estimating', 39: 'In Process',
    37: 'Ready for Sales', 38: 'Quotes Sent',
    27: 'Projected Deals', 34: 'Marketing Stalled', 30: 'Contact Made',
}
QUOTE_SENT_CUTOFF_YEAR = 2021
FAI_THRESHOLD = 0.50

HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
BODY_FONT = Font(name='Arial', size=10)
BORDER = Border(bottom=Side(style='thin', color='D9D9D9'))
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
BODY_ALIGN = Alignment(vertical='center')


BOOKING_HEADERS_DISPLAY = [
    'Database', 'Order Date', 'Sales Order No', 'So Status', 'Quote No',
    'Customer ID', 'Customer Name', 'SO line No', 'National Part ID',
    'Customer Part No', 'Part Description', 'Order Qty', 'Line Status',
    'Unit Price', 'Total Amount Ordered', 'Total Shipped Qty',
    'Ful Fill Qty', 'Pending Qty', 'Pending Amount'
]


def find_file(base_dir, pattern):
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if pattern in f:
                return os.path.join(root, f)
    return None


def load_org_ids(filepath):
    org_map = {}
    if not os.path.exists(filepath):
        print(f"  Warning: org_ids file not found: {filepath}")
        return org_map
    with open(filepath, 'r', encoding='utf-8-sig') as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            name = row.get('NAME', '').strip().upper()
            org_id = row.get('ORG ID', '').strip()
            if name and org_id:
                org_map[name] = org_id
    return org_map


def load_all_sheets(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip() if h else f'COL_{idx}' for idx, h in enumerate(row)]
                continue
            row_dict = {}
            for idx, h in enumerate(headers):
                val = row[idx] if idx < len(row) else None
                if hasattr(val, 'strftime'):
                    val = val.strftime('%Y-%m-%d')
                elif val is not None:
                    val = str(val)
                else:
                    val = ''
                row_dict[h] = val
            rows.append(row_dict)
        result[sheet_name] = {'headers': headers or [], 'rows': rows}
    wb.close()
    return result


def generate_natman_bookings(dev_booking_path, output_dir):
    print("  Reading Development Booking file...")
    wb = openpyxl.load_workbook(dev_booking_path, read_only=True, data_only=True)
    ws = wb.active
    raw_headers = None
    raw_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            raw_headers = list(row)
            continue
        raw_rows.append(list(row))
    wb.close()
    print(f"  Loaded {len(raw_rows):,} booking rows")

    hi = {str(h).strip(): idx for idx, h in enumerate(raw_headers) if h}

    def get_val(row, col_name, fallback_idx=None):
        for key in hi:
            if key.replace('_', ' ').upper() == col_name.replace('_', ' ').upper():
                return row[hi[key]] if hi[key] < len(row) else None
            if key.replace(' ', '_').upper() == col_name.replace(' ', '_').upper():
                return row[hi[key]] if hi[key] < len(row) else None
        if fallback_idx is not None and fallback_idx < len(row):
            return row[fallback_idx]
        return None

    col_map = {}
    expected_cols = [
        'Database', 'Order_Date', 'Sales_Order_No', 'So_Status', 'Quote_No',
        'Customer_ID', 'Customer_Name', 'SO_line_No', 'National_Part_ID',
        'Customer_Part_No', 'Part_Description', 'Order_Qty', 'Line_Status',
        'Unit_Price', 'Total_Amount_Ordered', 'Total_Shipped_Qty',
        'FulFill_Qty', 'Pending_Qty', 'Pending_Amount'
    ]
    for idx, col in enumerate(expected_cols):
        for key in hi:
            norm_key = key.replace('_', '').replace(' ', '').upper()
            norm_col = col.replace('_', '').replace(' ', '').upper()
            if norm_key == norm_col:
                col_map[col] = hi[key]
                break
        if col not in col_map and idx < len(raw_headers):
            col_map[col] = idx

    def extract_row(row):
        result = []
        for col in expected_cols:
            idx = col_map.get(col, None)
            val = row[idx] if idx is not None and idx < len(row) else None
            result.append(val)
        return result

    main_rows = [extract_row(r) for r in raw_rows]

    nat_part_idx = expected_cols.index('National_Part_ID')
    cust_part_idx = expected_cols.index('Customer_Part_No')
    order_date_idx = expected_cols.index('Order_Date')
    so_no_idx = expected_cols.index('Sales_Order_No')
    quote_no_idx = expected_cols.index('Quote_No')
    line_status_idx = expected_cols.index('Line_Status')
    db_idx = expected_cols.index('Database')
    so_status_idx = expected_cols.index('So_Status')
    cust_id_idx = expected_cols.index('Customer_ID')
    cust_name_idx = expected_cols.index('Customer_Name')
    so_line_idx = expected_cols.index('SO_line_No')
    part_desc_idx = expected_cols.index('Part_Description')
    order_qty_idx = expected_cols.index('Order_Qty')
    unit_price_idx = expected_cols.index('Unit_Price')
    total_amt_idx = expected_cols.index('Total_Amount_Ordered')
    shipped_idx = expected_cols.index('Total_Shipped_Qty')
    fulfill_idx = expected_cols.index('FulFill_Qty')
    pending_qty_idx = expected_cols.index('Pending_Qty')
    pending_amt_idx = expected_cols.index('Pending_Amount')

    unique_parts = {}
    for r in main_rows:
        cust_part = str(r[cust_part_idx]).strip() if r[cust_part_idx] else ''
        if cust_part and cust_part not in unique_parts:
            unique_parts[cust_part] = r

    unique_rows = []
    for cp, r in unique_parts.items():
        unique_rows.append([
            r[db_idx], r[line_status_idx], r[order_date_idx], r[so_no_idx],
            r[so_status_idx], r[quote_no_idx], r[cust_id_idx], r[cust_name_idx],
            r[so_line_idx], r[nat_part_idx], r[cust_part_idx], r[part_desc_idx]
        ])

    so_groups = defaultdict(lambda: {
        'order_date': None, 'so_no': None, 'quote_no': None,
        'cust_name': None, 'nat_part': None, 'cust_part': None,
        'order_qty': 0, 'total_amt': 0, 'shipped': 0,
        'fulfill': 0, 'pending_qty': 0, 'pending_amt': 0
    })
    for r in main_rows:
        so_key = str(r[so_no_idx])
        grp = so_groups[so_key]
        if grp['order_date'] is None:
            grp['order_date'] = r[order_date_idx]
            grp['so_no'] = r[so_no_idx]
            grp['quote_no'] = r[quote_no_idx]
            grp['cust_name'] = r[cust_name_idx]
            grp['nat_part'] = r[nat_part_idx]
            grp['cust_part'] = r[cust_part_idx]
        try:
            grp['order_qty'] += float(r[order_qty_idx] or 0)
        except (TypeError, ValueError):
            pass
        try:
            val = r[total_amt_idx]
            if val and str(val).strip() != '-':
                grp['total_amt'] += float(val)
        except (TypeError, ValueError):
            pass
        try:
            grp['shipped'] += float(r[shipped_idx] or 0)
        except (TypeError, ValueError):
            pass
        try:
            grp['fulfill'] += float(r[fulfill_idx] or 0)
        except (TypeError, ValueError):
            pass
        try:
            val = r[pending_qty_idx]
            if val and str(val).strip() != '-':
                grp['pending_qty'] += float(val)
        except (TypeError, ValueError):
            pass
        try:
            val = r[pending_amt_idx]
            if val and str(val).strip() != '-':
                grp['pending_amt'] += float(val)
        except (TypeError, ValueError):
            pass

    totals_rows = []
    for so_key, grp in so_groups.items():
        totals_rows.append([
            grp['order_date'], grp['so_no'], grp['quote_no'],
            grp['cust_name'], grp['nat_part'], grp['cust_part'],
            grp['order_qty'], grp['total_amt'], grp['shipped'],
            grp['fulfill'], grp['pending_qty'], grp['pending_amt']
        ])

    landmark_data = {}
    for r in main_rows:
        cust_part = str(r[cust_part_idx]).strip().upper() if r[cust_part_idx] else ''
        if not cust_part:
            continue
        order_date = r[order_date_idx]
        if not order_date or not hasattr(order_date, 'strftime'):
            try:
                if order_date and str(order_date).strip():
                    order_date = datetime.strptime(str(order_date).strip()[:10], '%Y-%m-%d')
                else:
                    continue
            except (ValueError, TypeError):
                continue

        if cust_part not in landmark_data:
            landmark_data[cust_part] = {
                'date': order_date,
                'nat_part': r[nat_part_idx],
                'order_no': r[so_no_idx],
                'quote_no': r[quote_no_idx],
                'last_date': order_date,
                'last_order_no': r[so_no_idx],
                'last_quote_no': r[quote_no_idx],
            }
        else:
            entry = landmark_data[cust_part]
            if order_date < entry['date']:
                entry['date'] = order_date
                entry['nat_part'] = r[nat_part_idx]
                entry['order_no'] = r[so_no_idx]
                entry['quote_no'] = r[quote_no_idx]
            if order_date > entry['last_date']:
                entry['last_date'] = order_date
                entry['last_order_no'] = r[so_no_idx]
                entry['last_quote_no'] = r[quote_no_idx]

    landmark_rows = []
    for cust_part, info in sorted(landmark_data.items()):
        landmark_rows.append([
            info['nat_part'], cust_part,
            info['date'], info['order_no'], info['quote_no'],
            info['last_date'], info['last_order_no'], info['last_quote_no'],
        ])

    print(f"  Generated: MAIN={len(main_rows):,}, UNIQUE={len(unique_rows):,}, "
          f"TOTALS={len(totals_rows):,}, LANDMARK={len(landmark_rows):,}")

    natman_path = os.path.join(output_dir, f'Natman_Bookings_{datetime.now().strftime("%Y%m%d")}.xlsx')
    wb_out = openpyxl.Workbook()

    def write_sheet_fast(ws, headers, data_rows):
        write_header(ws, headers)
        for row_data in data_rows:
            cleaned = [clean_val(v) for v in row_data]
            ws.append(cleaned)

    ws_main = wb_out.active
    ws_main.title = 'MAIN'
    write_sheet_fast(ws_main, BOOKING_HEADERS_DISPLAY, main_rows)

    ws_unique = wb_out.create_sheet('UNIQUE')
    unique_headers = ['Database', 'Line Status', 'Order Date', 'Sales Order No',
                      'So Status', 'Quote No', 'Customer ID', 'Customer Name',
                      'SO line No', 'National Part ID', 'Customer Part No', 'Part Description']
    write_sheet_fast(ws_unique, unique_headers, unique_rows)

    ws_totals = wb_out.create_sheet('TOTALS (By SO)')
    totals_headers = ['Order Date', 'Sales Order No', 'Quote No', 'Customer Name',
                      'NATIONAL PART ID', 'Customer Part No', 'Order Qty',
                      'Total Amount Ordered', 'Total Shipped Qty', 'Ful Fill Qty',
                      'Pending Qty', 'Pending Amount']
    write_sheet_fast(ws_totals, totals_headers, totals_rows)

    ws_landmark = wb_out.create_sheet('LANDMARK')
    landmark_headers = [
        'NAT PART ID', 'CUST PART ID',
        'FIRST ORDER DATE', 'FIRST ORDER NO', 'FIRST ORDER QUOTE NO',
        'LAST ORDER DATE', 'LAST ORDER NUMBER', 'LAST ORDER QUOTE NO',
    ]
    write_sheet_fast(ws_landmark, landmark_headers, landmark_rows)

    wb_out.save(natman_path)
    print(f"  Saved Natman_Bookings: {natman_path}")

    landmark_map = {}
    for cust_part, info in landmark_data.items():
        landmark_map[cust_part] = {
            'first_order_date': info['date'],
            'first_order_no': info['order_no'],
            'quote_no': info['quote_no'],
        }

    return natman_path, landmark_map


def generate_pdsync(consolidated_rows, pd_cache, output_dir):
    pdsync_path = os.path.join(output_dir, f'National_PDSync_PDUploadPreview_{datetime.now().strftime("%Y%m%d")}.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PD Upload Preview'

    pdsync_headers = [
        'PD_ID', 'PD_Title', 'ORG_ID', 'NAME', 'CUSTOMER_PART_ID',
        'Current_PD_Stage', 'Phase_Label', 'Mapped_Status', 'Mapped_Probability',
        'Mapped_Med_Rev', 'Mapped_PD_P1_Time', 'Mapped_PD_P2_Time',
        'Mapped_PD_P3_Time', 'Mapped_PD_P4_Time', 'Mapped_PD_P5_Time',
        'Quote_Number', 'Would_Update', 'DRY_RUN', 'Skip_Reason'
    ]
    write_header(ws, pdsync_headers)

    for idx, cr in enumerate(consolidated_rows, 2):
        cust_part = str(cr['cust_part']).strip().upper() if cr['cust_part'] else ''
        pd_id = pd_cache.get(cust_part)

        if pd_id:
            skip_reason = ''
            would_update = 'Yes'
        else:
            pd_id = None
            skip_reason = 'No existing PD match; candidate for new deal creation'
            would_update = 'No'

        write_row(ws, idx, [
            pd_id, '', cr.get('org_id', ''), cr.get('name', ''),
            cr.get('cust_part', ''), '', cr.get('phase_label', ''),
            cr.get('mapped_status', ''), cr.get('mapped_probability', ''),
            cr.get('mapped_med_rev', ''),
            cr.get('mapped_pd_p1_time', ''), cr.get('mapped_pd_p2_time', ''),
            '', cr.get('mapped_pd_p4_time', ''), cr.get('mapped_pd_p5_time', ''),
            cr.get('quote_number', ''), would_update, 'Yes', skip_reason
        ])

    wb.save(pdsync_path)
    print(f"  Saved PDSync: {pdsync_path}")
    return pdsync_path


def transform_raw_data(input_file, org_id_map):
    print(f"  Reading raw data from {os.path.basename(input_file)}...")
    src = openpyxl.load_workbook(input_file, data_only=True, read_only=True)
    ws = src.active
    headers, rows = None, []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        rows.append(list(row))
    src.close()
    print(f"  Loaded {len(rows):,} rows")

    ci = {h: idx for idx, h in enumerate(headers)}
    ID, NAME, STATUS = ci['ID'], ci['NAME'], ci['STATUS']
    WIN_PROB = ci['WIN_PROBABILITY']
    CUST_PART, DESC = ci['CUSTOMER_PART_ID'], ci['DESCRIPTION']
    QTY, UNIT_PRICE = ci['QTY'], ci['UNIT_PRICE']
    LINE_NO2, PART_ID2 = ci['LINE_NO2'], ci['PART_ID2']
    QUOTE_DATE, CREATE_DATE = ci['QUOTE_DATE'], ci['CREATE_DATE']
    PRODUCT_CODE = ci['PRODUCT_CODE']

    def is_nre(r):
        prod = str(r[PRODUCT_CODE]).upper().strip() if r[PRODUCT_CODE] else ''
        desc = str(r[DESC]).upper() if r[DESC] else ''
        if prod == 'TOOL':
            return True
        for kw in ['TOOLING', 'ENGINEERING CHARGE', 'NRE', 'CRATING OF TOOL']:
            if kw in desc:
                return True
        return False

    print("  Filtering...")
    filtered = []
    skip_c, skip_wp, skip_r, skip_f = 0, 0, 0, 0
    for r in rows:
        id_val = str(r[ID]).strip() if r[ID] is not None else ''
        watch_part = str(r[CUST_PART])
        watch_empty_part = bool(watch_part == "None")
        watch_price = float(r[UNIT_PRICE]) if r[UNIT_PRICE] is not None else 0
        watch_fat = bool(re.match(pattern="(\\bFA[TI]\\b)", string=watch_part))

        if id_val.upper().startswith('C'):
            skip_c += 1
            continue
        wp = r[WIN_PROB]
        try:
            wp_float = float(wp) if wp is not None and str(wp).strip() != '' else None
        except (TypeError, ValueError):
            wp_float = None
        if wp_float == 0.8:
            skip_wp += 1
            continue
        if r[STATUS] == 'R':
            skip_r += 1
            continue
        if watch_fat or watch_empty_part:
            skip_f += 1
            continue
        if watch_price == 0:
            skip_f += 1
            continue
        r[ID] = id_val
        filtered.append(r)

    print(f"  Removed {skip_c:,} (C-prefix) + {skip_wp:,} (WinProb=0.8) + {skip_r:,} (Status=R) + {skip_f:,} (Misc)")

    nre_by_quote = defaultdict(float)
    part_rows = []
    nre_count = 0
    for r in filtered:
        if is_nre(r):
            price = r[UNIT_PRICE] if r[UNIT_PRICE] is not None else 0
            qty = r[QTY] if r[QTY] is not None else 0
            nre_by_quote[r[ID]] += qty * price
            nre_count += 1
        else:
            part_rows.append(r)
    print(f"  NRE lines separated: {nre_count:,}")

    print("  Selecting median price breaks...")
    groups = defaultdict(list)
    for r in part_rows:
        groups[(r[ID], str(r[LINE_NO2]), str(r[PART_ID2]))].append(r)

    result_rows = []
    for key, grp in groups.items():
        if len(grp) == 1:
            result_rows.append(grp[0])
            continue
        qtys = [(i, r[QTY]) for i, r in enumerate(grp) if r[QTY] is not None]
        if not qtys:
            result_rows.append(grp[0])
            continue
        med_qty = statistics.median([q for _, q in qtys])
        best_idx = min(qtys, key=lambda x: abs(x[1] - med_qty))[0]
        result_rows.append(grp[best_idx])

    print(f"  After median selection: {len(result_rows):,}")

    pre_cutoff_removed = 0
    filtered_result = []
    for r in result_rows:
        qd = r[QUOTE_DATE]
        if hasattr(qd, 'year') and qd.year < QUOTE_SENT_CUTOFF_YEAR:
            pre_cutoff_removed += 1
            continue
        filtered_result.append(r)
    result_rows = filtered_result
    print(f"  Removed {pre_cutoff_removed:,} rows with QUOTE_DATE before {QUOTE_SENT_CUTOFF_YEAR}")

    def map_status(row):
        s = row[STATUS]
        price = row[UNIT_PRICE] if row[UNIT_PRICE] is not None else 0
        m = {'A': 'RFQ - in-process', 'P': 'Quote sent', 'L': 'Lost', 'W': 'Won'}
        if s in m:
            return m[s]
        if s == 'X':
            return 'no quote' if price == 0 else 'dormant'
        return str(s)

    def map_prob(row):
        wp = row[WIN_PROB]
        try:
            wp_float = float(wp) if wp is not None and str(wp).strip() != '' else None
        except (TypeError, ValueError):
            wp_float = None
        if wp_float is not None:
            if wp_float == 0.7:
                return 'NP'
            if wp_float == 0.9:
                return 'NC'
        return ''

    def row_rev(r):
        qty = r[QTY] if r[QTY] is not None else 0
        price = r[UNIT_PRICE] if r[UNIT_PRICE] is not None else 0
        med_rev = qty * price
        has_nre = r[ID] in nre_by_quote
        nre_amt = nre_by_quote.get(r[ID], 0)
        return med_rev + (nre_amt if has_nre else 0)

    print(f"  Consolidating to one entry per CUSTOMER_PART_ID...")
    part_groups = defaultdict(list)
    for r in result_rows:
        cp = str(r[CUST_PART]).strip().upper() if r[CUST_PART] else ''
        if not cp:
            cp = f"_BLANK_{r[ID]}"
        part_groups[cp].append(r)

    print(f"  {len(result_rows):,} rows -> {len(part_groups):,} unique parts")

    consolidated_rows = []

    for cp_key, grp in part_groups.items():
        grp_with_rev = [(r, row_rev(r)) for r in grp]
        won_rows = [(r, rev) for r, rev in grp_with_rev if map_status(r) == 'Won']
        all_revs = [rev for _, rev in grp_with_rev if rev > 0]
        max_rev = max(all_revs) if all_revs else 0

        create_dates = [r[CREATE_DATE] for r, _ in grp_with_rev if hasattr(r[CREATE_DATE], 'strftime')]
        earliest_p1 = min(create_dates) if create_dates else None

        quote_dates = [(r, r[QUOTE_DATE]) for r, _ in grp_with_rev if hasattr(r[QUOTE_DATE], 'strftime')]
        if quote_dates:
            most_recent_row, latest_p2 = max(quote_dates, key=lambda x: x[1])
        else:
            most_recent_row = grp[0]
            latest_p2 = None

        rep = most_recent_row
        highest_rev_row, highest_rev = max(grp_with_rev, key=lambda x: x[1]) if grp_with_rev else (grp[0], 0)
        prob_val = map_prob(most_recent_row)

        def make_consolidated(status, value, p1_dt, p2_dt, p4_dt, p5_dt, quote_row, phase_label):
            p1_str = p1_dt.strftime('%Y-%m-%d') if hasattr(p1_dt, 'strftime') else ''
            p2_str = p2_dt.strftime('%Y-%m-%d') if hasattr(p2_dt, 'strftime') else ''
            p4_str = p4_dt.strftime('%Y-%m-%d') if hasattr(p4_dt, 'strftime') else (p4_dt if isinstance(p4_dt, str) else '')
            p5_str = p5_dt.strftime('%Y-%m-%d') if hasattr(p5_dt, 'strftime') else (p5_dt if isinstance(p5_dt, str) else '')
            quote_id = str(quote_row[ID]).strip() if quote_row[ID] is not None else ''
            name_upper = str(rep[NAME]).strip().upper() if rep[NAME] else ''
            return {
                'id': quote_row[ID],
                'org_id': org_id_map.get(name_upper, ''),
                'name': rep[NAME],
                'status': status,
                'prob': prob_val,
                'cust_part': rep[CUST_PART],
                'description': rep[DESC],
                'med_rev': round(value, 2),
                'quote_number': quote_id,
                'p1_str': p1_str,
                'p2_str': p2_str,
                'p4_str': p4_str,
                'p5_str': p5_str,
                'phase_label': phase_label,
                'mapped_status': status,
                'mapped_probability': prob_val,
                'mapped_med_rev': round(value, 2),
                'mapped_pd_p1_time': p1_str,
                'mapped_pd_p2_time': p2_str,
                'mapped_pd_p4_time': p4_str,
                'mapped_pd_p5_time': p5_str,
            }

        if won_rows:
            won_row, won_rev = max(won_rows, key=lambda x: x[1])
            won_p2 = won_row[QUOTE_DATE]

            if max_rev > 0 and won_rev < (FAI_THRESHOLD * max_rev):
                consolidated_rows.append(make_consolidated(
                    status='Won', value=won_rev, p1_dt=earliest_p1, p2_dt=won_p2,
                    p4_dt='', p5_dt=won_p2, quote_row=won_row, phase_label='Phase 5'
                ))
                non_won = [(r, rev) for r, rev in grp_with_rev if map_status(r) != 'Won']
                if non_won:
                    best_opp_row, best_opp_rev = max(non_won, key=lambda x: x[1])
                else:
                    best_opp_row, best_opp_rev = highest_rev_row, highest_rev
                consolidated_rows.append(make_consolidated(
                    status='Phase 4 - Production Opp', value=best_opp_rev,
                    p1_dt=earliest_p1, p2_dt=latest_p2, p4_dt=won_p2,
                    p5_dt='', quote_row=best_opp_row, phase_label='Phase 4'
                ))
            else:
                consolidated_rows.append(make_consolidated(
                    status='Won', value=max(won_rev, max_rev), p1_dt=earliest_p1,
                    p2_dt=latest_p2, p4_dt='', p5_dt=won_p2,
                    quote_row=won_row, phase_label='Phase 5'
                ))
        else:
            consolidated_rows.append(make_consolidated(
                status=map_status(most_recent_row), value=max_rev,
                p1_dt=earliest_p1, p2_dt=latest_p2, p4_dt='', p5_dt='',
                quote_row=most_recent_row if max_rev == 0 else highest_rev_row,
                phase_label=''
            ))

    print(f"  Consolidated: {len(consolidated_rows):,} rows")
    return consolidated_rows


def match_with_pd_cache(consolidated_rows, pd_cache):
    matched_rows = []
    unmatched_rows = []

    for cr in consolidated_rows:
        cust_part = str(cr['cust_part']).strip().upper() if cr['cust_part'] else ''
        pd_id = pd_cache.get(cust_part)
        if pd_id is not None:
            cr['pd_id'] = pd_id
            matched_rows.append(cr)
        else:
            cr['pd_id'] = ''
            unmatched_rows.append(cr)

    print(f"  PD matched: {len(matched_rows):,} | Unmatched: {len(unmatched_rows):,}")
    return matched_rows, unmatched_rows


def translate_field(field_key, raw_value):
    if raw_value is None or raw_value == '':
        return ''
    field_opts = OPTION_LABELS.get(field_key, {})
    def lookup(value):
        if value in field_opts:
            return field_opts[value]
        value_str = str(value).strip()
        if value_str in field_opts:
            return field_opts[value_str]
        try:
            value_int = int(value_str)
            if value_int in field_opts:
                return field_opts[value_int]
        except (TypeError, ValueError):
            pass
        return value_str
    if isinstance(raw_value, list):
        return ', '.join(str(lookup(v)) for v in raw_value)
    raw_str = str(raw_value)
    if ',' in raw_str:
        parts = [p.strip() for p in raw_str.split(',') if p.strip()]
        return ', '.join(str(lookup(p)) for p in parts)
    return str(lookup(raw_value))


def fetch_deal_details(deal_ids):
    if not API_TOKEN:
        print("  WARNING: No PIPEDRIVE_API_KEY set, skipping deal detail fetch")
        return {}

    print(f"  Fetching details for {len(deal_ids)} Pipedrive deals...")
    deal_details = {}
    unique_ids = list(set(deal_ids))
    fetched = 0
    errors = 0

    for deal_id in unique_ids:
        try:
            resp = requests.get(
                f'{BASE_URL}/deals/{deal_id}',
                params={'api_token': API_TOKEN},
                timeout=30
            )
            if resp.status_code == 429:
                print("  Rate limited, waiting 10s...")
                time.sleep(10)
                resp = requests.get(
                    f'{BASE_URL}/deals/{deal_id}',
                    params={'api_token': API_TOKEN},
                    timeout=30
                )
            if resp.status_code == 200:
                deal = resp.json().get('data', {})
                if deal:
                    org = deal.get('org_id') or {}
                    raw_stage = deal.get('stage_id', '')
                    stage_label = PHASE_IDS.get(raw_stage, str(raw_stage) if raw_stage else '')
                    deal_details[deal_id] = {
                        'title': deal.get('title', ''),
                        'value': deal.get('value', ''),
                        'status': deal.get('status', ''),
                        'won_time': deal.get('won_time', ''),
                        'org_name': org.get('name', '') if isinstance(org, dict) else '',
                        'org_id': org.get('value', '') if isinstance(org, dict) else '',
                        'stage_id': stage_label,
                        'label': translate_field('label', deal.get('label', '')),
                        'platform_company': translate_field(PD_FIELDS['platform_company'], deal.get(PD_FIELDS['platform_company'], '')),
                        'deal_type': translate_field(PD_FIELDS['deal_type'], deal.get(PD_FIELDS['deal_type'], '')),
                        'mfg_type': translate_field(PD_FIELDS['mfg_type'], deal.get(PD_FIELDS['mfg_type'], '')),
                        'industry': translate_field(PD_FIELDS['industry'], deal.get(PD_FIELDS['industry'], '')),
                        'quote_number': deal.get(PD_FIELDS['quote_number'], ''),
                        'po_number': deal.get(PD_FIELDS['po_number'], ''),
                        'p1_time': deal.get(PD_FIELDS['p1_time'], ''),
                        'p2_time': deal.get(PD_FIELDS['p2_time'], ''),
                        'p3_time': deal.get(PD_FIELDS['p3_time'], ''),
                        'p4_time': deal.get(PD_FIELDS['p4_time'], ''),
                        'p5_time': deal.get(PD_FIELDS['p5_time'], ''),
                        'part': deal.get(PD_FIELDS['part'], ''),
                        'customer_part': deal.get(PD_FIELDS['customer_part'], ''),
                    }
                    fetched += 1
            else:
                errors += 1
        except Exception as e:
            errors += 1
            if errors <= 3:
                print(f"  Error fetching deal {deal_id}: {e}")

        if fetched % 25 == 0 and fetched > 0:
            print(f"  Fetched {fetched}/{len(unique_ids)} deals...", end='\r')

        if fetched % 100 == 0 and fetched > 0:
            time.sleep(0.5)

    print(f"\n  Fetched {fetched} deal details ({errors} errors)")
    return deal_details


def clean_val(v):
    if isinstance(v, str):
        return v.strip()
    if hasattr(v, 'strftime'):
        return v.strftime('%m/%d/%Y')
    return v


def write_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER_ALIGN


def write_row(ws, row_idx, values, formats=None):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=c, value=clean_val(val))
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = BODY_ALIGN
        if formats and c <= len(formats) and formats[c-1]:
            cell.number_format = formats[c-1]


def set_col_widths(ws, widths):
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def search_pipedrive_deals(customer_parts):
    if not API_TOKEN:
        print("  WARNING: No PIPEDRIVE_API_KEY set, skipping Pipedrive search")
        return {}

    print(f"  Searching Pipedrive for {len(customer_parts)} unique customer parts...")
    pd_cache = {}
    searched = 0
    found = 0
    errors = 0

    for cust_part in customer_parts:
        if not cust_part or cust_part == 'NONE':
            continue
        try:
            resp = requests.get(
                f'{BASE_URL}/deals/search',
                params={
                    'api_token': API_TOKEN,
                    'term': cust_part,
                    'fields': 'custom_fields',
                    'limit': 1,
                },
                timeout=30
            )
            if resp.status_code == 429:
                print("  Rate limited, waiting 10s...")
                time.sleep(10)
                resp = requests.get(
                    f'{BASE_URL}/deals/search',
                    params={
                        'api_token': API_TOKEN,
                        'term': cust_part,
                        'fields': 'custom_fields',
                        'limit': 1,
                    },
                    timeout=30
                )
            if resp.status_code == 200:
                data = resp.json().get('data', {})
                items = data.get('items', []) if data else []
                for item in (items or []):
                    deal_item = item.get('item', {})
                    deal_id = deal_item.get('id')
                    if not deal_id:
                        continue
                    try:
                        detail_resp = requests.get(
                            f'{BASE_URL}/deals/{deal_id}',
                            params={'api_token': API_TOKEN},
                            timeout=30
                        )
                        if detail_resp.status_code == 200:
                            deal_data = detail_resp.json().get('data', {})
                            deal_cust_part = str(deal_data.get(PD_FIELDS['customer_part'], '') or '').strip().upper()
                            if deal_cust_part == cust_part:
                                pd_cache[cust_part] = deal_id
                                found += 1
                                break
                    except Exception:
                        pass
            else:
                errors += 1

            searched += 1
            if searched % 50 == 0:
                print(f"  Searched {searched}/{len(customer_parts)} parts ({found} found)...")
                time.sleep(0.5)
            elif searched % 10 == 0:
                time.sleep(0.2)

        except Exception as e:
            errors += 1
            if errors <= 3:
                print(f"  Error searching for {cust_part}: {e}")

    print(f"  Pipedrive search complete: {found} matches found out of {searched} searched ({errors} errors)")
    return pd_cache


def main(cli_args=None):
    global OUTPUT_DIR, OUTPUT_FILE, QUOTE_SENT_CUTOFF_YEAR, FAI_THRESHOLD

    parser = argparse.ArgumentParser(description='Parts Analysis')
    parser.add_argument('--national-file', help='Path to National QuoteData xlsx')
    parser.add_argument('--booking-file', help='Path to Development Booking xlsx')
    parser.add_argument('--output-dir', help='Output directory')
    parser.add_argument('--cutoff-year', type=int, help='Quote date cutoff year')
    parser.add_argument('--fai-threshold', type=float, help='FAI threshold (0-1)')
    parser.add_argument('--json-output', help='Path to write JSON summary')
    parser.add_argument('--pd-cache-file', help='Path to pd_cache.json (optional)')
    parser.add_argument('--previous-run-json', help='Path to JSON with previous run parts for diff')
    args = parser.parse_args(cli_args)

    if args.output_dir:
        OUTPUT_DIR = args.output_dir
        OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'Parts_Analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    if args.cutoff_year:
        QUOTE_SENT_CUTOFF_YEAR = args.cutoff_year
    if args.fai_threshold is not None:
        FAI_THRESHOLD = args.fai_threshold

    national_file = args.national_file
    booking_file = args.booking_file

    if not national_file or not os.path.exists(national_file):
        print("  ERROR: National QuoteData file is required (--national-file)")
        sys.exit(1)
    if not booking_file or not os.path.exists(booking_file):
        print("  ERROR: Development Booking file is required (--booking-file)")
        sys.exit(1)

    start_time = time.time()
    print("=" * 60)
    print("  PARTS ANALYSIS — National QuoteData + Development Booking")
    print("=" * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    prev_new_deals = {}
    prev_pd_info = {}
    if args.previous_run_json and os.path.exists(args.previous_run_json):
        try:
            with open(args.previous_run_json, 'r') as pf:
                prev_data = json.load(pf)
            for p in prev_data:
                key = str(p.get('customerPartId', '')).strip().upper()
                if not key:
                    continue
                if p.get('sheetType') == 'new_deals':
                    prev_new_deals[key] = p
                elif p.get('sheetType') == 'pd_info':
                    prev_pd_info[key] = p
            print(f"  Loaded previous run: {len(prev_new_deals)} new_deals, {len(prev_pd_info)} pd_info parts")
        except Exception as e:
            print(f"  Warning: Could not load previous run data: {e}")

    print(f"\n[1/7] Processing Development Booking...")
    print(f"  Input: {os.path.basename(booking_file)}")
    natman_path, landmark = generate_natman_bookings(booking_file, OUTPUT_DIR)

    print(f"\n[2/7] Loading reference data...")
    bundled_org_ids = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'org_ids.csv')
    org_id_map = load_org_ids(bundled_org_ids) if os.path.exists(bundled_org_ids) else {}
    print(f"  Loaded {len(org_id_map)} org ID mappings")

    print(f"\n[3/7] Transforming raw quote data...")
    print(f"  Input: {os.path.basename(national_file)}")
    consolidated_rows = transform_raw_data(national_file, org_id_map)

    print(f"\n[4/7] Matching with Pipedrive...")
    pd_cache = {}
    pd_cache_file = args.pd_cache_file
    if not pd_cache_file:
        bundled_cache = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pd_cache.json')
        if os.path.exists(bundled_cache):
            pd_cache_file = bundled_cache

    if pd_cache_file and os.path.exists(pd_cache_file):
        with open(pd_cache_file, 'r') as f:
            raw_cache = json.load(f)
        pd_cache = {k.strip().upper(): v for k, v in raw_cache.items() if v is not None}
        print(f"  Loaded {len(pd_cache)} PD cache entries from {os.path.basename(pd_cache_file)}")
    else:
        unique_parts = set()
        for cr in consolidated_rows:
            cp = str(cr['cust_part']).strip().upper() if cr['cust_part'] else ''
            if cp and not cp.startswith('_BLANK_'):
                unique_parts.add(cp)
        pd_cache = search_pipedrive_deals(sorted(unique_parts))

    matched_rows, unmatched_rows = match_with_pd_cache(consolidated_rows, pd_cache)

    print(f"\n[5/7] Generating PDSync output...")
    pdsync_path = generate_pdsync(consolidated_rows, pd_cache, OUTPUT_DIR)

    print(f"\n[6/7] Pipedrive deal detail enrichment...")
    deal_details = {}
    deal_ids = [r['pd_id'] for r in matched_rows if r['pd_id']]

    deals_export_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pd_deals_export.json')
    if os.path.exists(deals_export_file):
        with open(deals_export_file, 'r') as f:
            all_deals = json.load(f)
        for did in deal_ids:
            dd = all_deals.get(str(did))
            if dd:
                deal_details[did] = dd
        print(f"  Loaded {len(deal_details)}/{len(deal_ids)} deal details from local export ({len(all_deals)} total deals)")
    elif deal_ids and API_TOKEN and os.environ.get('FETCH_PD_DETAILS', '').lower() == 'true':
        deal_details = fetch_deal_details(deal_ids)
    else:
        print(f"  No deals export found and live API fetch disabled ({len(deal_ids)} deals matched)")
        print(f"  Place pd_deals_export.json in scripts/src/ or set FETCH_PD_DETAILS=true")

    print(f"\n[7/7] Writing output spreadsheet...")
    wb = openpyxl.Workbook()

    ws_parts = wb.active
    ws_parts.title = "All_Unique_Parts"
    all_parts = set()
    for cr in consolidated_rows:
        cp = str(cr['cust_part']).strip() if cr['cust_part'] else ''
        if cp:
            all_parts.add(cp)
    for entry in landmark:
        all_parts.add(entry)

    all_parts_sorted = sorted(all_parts)
    write_header(ws_parts, ['CUSTOMER_PART_ID', 'IN_QUOTE_DATA', 'IN_LANDMARK', 'HAS_PD_MATCH'])

    quote_parts = {str(cr['cust_part']).strip().upper() for cr in consolidated_rows if cr['cust_part']}
    for idx, part in enumerate(all_parts_sorted, 2):
        part_upper = part.strip().upper()
        in_quote = 'Yes' if part_upper in quote_parts else 'No'
        in_landmark = 'Yes' if part_upper in landmark else 'No'
        has_pd = 'Yes' if part_upper in pd_cache else 'No'
        write_row(ws_parts, idx, [part, in_quote, in_landmark, has_pd])

    set_col_widths(ws_parts, [25, 15, 15, 15])
    ws_parts.freeze_panes = 'A2'
    ws_parts.auto_filter.ref = f"A1:D{len(all_parts_sorted) + 1}"
    print(f"  All_Unique_Parts: {len(all_parts_sorted)} parts")

    ws_new = wb.create_sheet("New_Deals")
    has_diff = True
    new_headers = [
        'ORG_ID', 'NAME', 'CUSTOMER_PART_ID',
        'Mapped_Status', 'Mapped_Probability', 'Mapped_Med_Rev',
        'Mapped_PD_P1_Time', 'Mapped_PD_P2_Time',
        'Mapped_PD_P4_Time', 'Mapped_PD_P5_Time',
        'Quote_Number',
        'FIRST_ORDER_DATE', 'FIRST_ORDER_NO', 'LANDMARK_QUOTE_NO',
        'CALC_LABEL'
    ]
    if has_diff:
        new_headers.append('Change')
    write_header(ws_new, new_headers)
    new_formats = [None, None, None, None, None, '$#,##0.00',
                   'MM/DD/YYYY', 'MM/DD/YYYY', 'MM/DD/YYYY', 'MM/DD/YYYY',
                   None, 'MM/DD/YYYY', None, None, None]
    if has_diff:
        new_formats.append(None)

    for idx, cr in enumerate(unmatched_rows, 2):
        cust_part = str(cr['cust_part']).strip().upper() if cr['cust_part'] else ''
        lm = landmark.get(cust_part, {})
        first_order_date = lm.get('first_order_date', '')
        first_order_no = lm.get('first_order_no', '')
        lm_quote_no = lm.get('quote_no', '')

        mapped_prob = str(cr.get('mapped_probability', '')).strip().upper()
        if mapped_prob == 'NC':
            calc_label = 'New Customer'
        else:
            calc_label = 'New Part'
            p2_str = cr.get('mapped_pd_p2_time', '')
            if first_order_date and p2_str:
                try:
                    if hasattr(first_order_date, 'strftime'):
                        fod = first_order_date
                    else:
                        fod = datetime.strptime(str(first_order_date), '%Y-%m-%d')

                    if isinstance(p2_str, str) and p2_str:
                        p2_date = datetime.strptime(p2_str, '%Y-%m-%d')
                    elif hasattr(p2_str, 'strftime'):
                        p2_date = p2_str
                    else:
                        p2_date = None

                    if p2_date and fod and p2_date > fod:
                        calc_label = 'Repeat'
                except (ValueError, TypeError):
                    pass

        row_vals = [
            cr.get('org_id', ''), cr.get('name', ''), cr.get('cust_part', ''),
            cr.get('mapped_status', ''), cr.get('mapped_probability', ''),
            cr.get('mapped_med_rev', ''),
            cr.get('mapped_pd_p1_time', ''), cr.get('mapped_pd_p2_time', ''),
            cr.get('mapped_pd_p4_time', ''), cr.get('mapped_pd_p5_time', ''),
            cr.get('quote_number', ''),
            first_order_date, first_order_no, lm_quote_no,
            calc_label
        ]
        if has_diff:
            if not prev_new_deals:
                change_val = ''
            elif cust_part not in prev_new_deals:
                change_val = 'NEW'
            else:
                prev = prev_new_deals[cust_part]
                compare_fields = [
                    ('mappedStatus', cr.get('mapped_status', '')),
                    ('mappedProbability', cr.get('mapped_probability', '')),
                    ('mappedMedRev', cr.get('mapped_med_rev', '')),
                ]
                changed = False
                for pkey, curr_val in compare_fields:
                    prev_val = prev.get(pkey, '')
                    if str(curr_val).strip() != str(prev_val).strip():
                        changed = True
                        break
                change_val = 'CHANGED' if changed else ''
            row_vals.append(change_val)
        write_row(ws_new, idx, row_vals, new_formats)

    nd_col_count = len(new_headers)
    nd_col_letter = get_column_letter(nd_col_count)
    col_widths_new = [10, 35, 22, 20, 18, 16, 14, 14, 14, 14, 14, 14, 14, 14, 12]
    if has_diff:
        col_widths_new.append(12)
    set_col_widths(ws_new, col_widths_new)
    ws_new.freeze_panes = 'A2'
    ws_new.auto_filter.ref = f"A1:{nd_col_letter}{len(unmatched_rows) + 1}"
    repeat_count = sum(1 for r in range(2, len(unmatched_rows) + 2) if ws_new.cell(row=r, column=15).value == 'Repeat')
    new_count = sum(1 for r in range(2, len(unmatched_rows) + 2) if ws_new.cell(row=r, column=15).value == 'New Part')
    nc_count = sum(1 for r in range(2, len(unmatched_rows) + 2) if ws_new.cell(row=r, column=15).value == 'New Customer')
    print(f"  New_Deals: {len(unmatched_rows)} rows ({repeat_count} Repeat, {new_count} New Part, {nc_count} New Customer)")

    ws_pd = wb.create_sheet("PD_Info")
    pd_headers = [
        'PD_ID', 'ORG_ID', 'NAME', 'CUSTOMER_PART_ID',
        'Mapped_Status', 'Mapped_Probability', 'Mapped_Med_Rev',
        'Mapped_PD_P1_Time', 'Mapped_PD_P2_Time',
        'Mapped_PD_P4_Time', 'Mapped_PD_P5_Time',
        'Quote_Number',
        'FIRST_ORDER_DATE', 'FIRST_ORDER_NO', 'LANDMARK_QUOTE_NO',
        'PD_title', 'PD_value', 'PD_status', 'PD_stage',
        'PD_label', 'PD_industry', 'PD_deal_type', 'PD_mfg_type',
    ]
    if has_diff:
        pd_headers.append('Change')
    write_header(ws_pd, pd_headers)

    for idx, cr in enumerate(matched_rows, 2):
        pd_id = cr.get('pd_id', '')
        dd = deal_details.get(pd_id, {})
        cust_part = str(cr['cust_part']).strip().upper() if cr['cust_part'] else ''
        lm = landmark.get(cust_part, {})
        pd_row_vals = [
            pd_id, cr.get('org_id', ''), cr.get('name', ''), cr.get('cust_part', ''),
            cr.get('mapped_status', ''), cr.get('mapped_probability', ''),
            cr.get('mapped_med_rev', ''),
            cr.get('mapped_pd_p1_time', ''), cr.get('mapped_pd_p2_time', ''),
            cr.get('mapped_pd_p4_time', ''), cr.get('mapped_pd_p5_time', ''),
            cr.get('quote_number', ''),
            lm.get('first_order_date', ''), lm.get('first_order_no', ''), lm.get('quote_no', ''),
            dd.get('title', ''), dd.get('value', ''),
            dd.get('status', ''), dd.get('stage_id', ''),
            dd.get('label', ''), dd.get('industry', ''),
            dd.get('deal_type', ''), dd.get('mfg_type', ''),
        ]
        if has_diff:
            if not prev_pd_info:
                pd_change = ''
            elif cust_part not in prev_pd_info:
                pd_change = 'NEW'
            else:
                prev = prev_pd_info[cust_part]
                pd_compare = [
                    ('mappedStatus', cr.get('mapped_status', '')),
                    ('mappedProbability', cr.get('mapped_probability', '')),
                    ('mappedMedRev', cr.get('mapped_med_rev', '')),
                    ('pdValue', dd.get('value', '')),
                    ('pdStatus', dd.get('status', '')),
                    ('pdStage', dd.get('stage_id', '')),
                ]
                pd_changed = False
                for pkey, curr_val in pd_compare:
                    prev_val = prev.get(pkey, '')
                    if str(curr_val).strip() != str(prev_val).strip():
                        pd_changed = True
                        break
                pd_change = 'CHANGED' if pd_changed else ''
            pd_row_vals.append(pd_change)
        write_row(ws_pd, idx, pd_row_vals)

    pd_col_count = len(pd_headers)
    pd_col_letter = get_column_letter(pd_col_count)
    col_widths_pd = [10, 22, 30, 12, 10, 20, 30, 10, 10, 18, 18, 18, 18, 18, 14, 14, 12, 12, 12, 12, 12, 20, 20]
    if has_diff:
        col_widths_pd.append(12)
    set_col_widths(ws_pd, col_widths_pd)
    ws_pd.freeze_panes = 'A2'
    ws_pd.auto_filter.ref = f"A1:{pd_col_letter}{len(matched_rows) + 1}"
    print(f"  PD_Info: {len(matched_rows)} rows ({len(deal_details)} deals fetched from Pipedrive)")

    wb.save(OUTPUT_FILE)
    elapsed = time.time() - start_time
    print(f"\n{'=' * 60}")
    print(f"  DONE in {elapsed:.1f}s")
    print(f"  Output: {OUTPUT_FILE}")
    print(f"{'=' * 60}")

    MAX_SOURCE_ROWS = 5000
    print("  Loading generated output files for dashboard viewing...")
    natman_sheets = load_all_sheets(natman_path)
    for sn, sd in natman_sheets.items():
        total = len(sd['rows'])
        if total > MAX_SOURCE_ROWS:
            sd['rows'] = sd['rows'][:MAX_SOURCE_ROWS]
            print(f"    Natman_Bookings/{sn}: {total:,} rows (capped to {MAX_SOURCE_ROWS:,} for display)")
        else:
            print(f"    Natman_Bookings/{sn}: {total:,} rows")
    pdsync_sheets = load_all_sheets(pdsync_path)
    for sn, sd in pdsync_sheets.items():
        print(f"    PDSync/{sn}: {len(sd['rows']):,} rows")

    new_deals_data = []
    for cr in unmatched_rows:
        cust_part = str(cr['cust_part']).strip().upper() if cr['cust_part'] else ''
        lm = landmark.get(cust_part, {})
        mapped_prob = str(cr.get('mapped_probability', '')).strip().upper()
        if mapped_prob == 'NC':
            calc_label = 'New Customer'
        else:
            calc_label = 'New Part'
            p2_str = cr.get('mapped_pd_p2_time', '')
            fod_raw = lm.get('first_order_date', '')
            if fod_raw and p2_str:
                try:
                    fod = fod_raw if hasattr(fod_raw, 'strftime') else datetime.strptime(str(fod_raw), '%Y-%m-%d')
                    p2_date = datetime.strptime(p2_str, '%Y-%m-%d') if isinstance(p2_str, str) else p2_str
                    if p2_date and fod and p2_date > fod:
                        calc_label = 'Repeat'
                except (ValueError, TypeError):
                    pass
        new_deals_data.append({
            'org_id': str(cr.get('org_id', '')),
            'name': str(cr.get('name', '')),
            'customer_part_id': str(cr.get('cust_part', '')),
            'mapped_status': str(cr.get('mapped_status', '')),
            'mapped_probability': str(cr.get('mapped_probability', '')),
            'mapped_med_rev': float(cr.get('mapped_med_rev', 0) or 0),
            'mapped_pd_p1_time': str(cr.get('mapped_pd_p1_time', '')),
            'mapped_pd_p2_time': str(cr.get('mapped_pd_p2_time', '')),
            'mapped_pd_p4_time': str(cr.get('mapped_pd_p4_time', '')),
            'mapped_pd_p5_time': str(cr.get('mapped_pd_p5_time', '')),
            'quote_number': str(cr.get('quote_number', '')),
            'first_order_date': lm.get('first_order_date', '').strftime('%Y-%m-%d') if hasattr(lm.get('first_order_date', ''), 'strftime') else str(lm.get('first_order_date', '')),
            'first_order_no': str(lm.get('first_order_no', '')),
            'landmark_quote_no': str(lm.get('quote_no', '')),
            'calc_label': calc_label,
        })

    pd_info_data = []
    for cr in matched_rows:
        pd_id = cr.get('pd_id', '')
        dd = deal_details.get(pd_id, {})
        cust_part_upper = str(cr['cust_part']).strip().upper() if cr['cust_part'] else ''
        lm_pd = landmark.get(cust_part_upper, {})
        pd_info_data.append({
            'pd_id': pd_id,
            'org_id': str(cr.get('org_id', '')),
            'name': str(cr.get('name', '')),
            'customer_part_id': str(cr.get('cust_part', '')),
            'mapped_status': str(cr.get('mapped_status', '')),
            'mapped_probability': str(cr.get('mapped_probability', '')),
            'mapped_med_rev': float(cr.get('mapped_med_rev', 0) or 0),
            'mapped_pd_p1_time': str(cr.get('mapped_pd_p1_time', '')),
            'mapped_pd_p2_time': str(cr.get('mapped_pd_p2_time', '')),
            'mapped_pd_p4_time': str(cr.get('mapped_pd_p4_time', '')),
            'mapped_pd_p5_time': str(cr.get('mapped_pd_p5_time', '')),
            'quote_number': str(cr.get('quote_number', '')),
            'first_order_date': lm_pd.get('first_order_date', '').strftime('%Y-%m-%d') if hasattr(lm_pd.get('first_order_date', ''), 'strftime') else str(lm_pd.get('first_order_date', '')),
            'first_order_no': str(lm_pd.get('first_order_no', '')),
            'landmark_quote_no': str(lm_pd.get('quote_no', '')),
            'value': float(dd.get('value', 0) or 0),
            'status': str(dd.get('status', '')),
            'org_name': str(dd.get('org_name', '')),
            'stage_id': str(dd.get('stage_id', '')),
            'label': str(dd.get('label', '')),
            'platform_company': str(dd.get('platform_company', '')),
            'deal_type': str(dd.get('deal_type', '')),
            'mfg_type': str(dd.get('mfg_type', '')),
            'industry': str(dd.get('industry', '')),
            'pd_quote_number': str(dd.get('quote_number', '') or ''),
            'pd_po_number': str(dd.get('po_number', '') or ''),
        })

    all_parts_data = []
    for part in all_parts_sorted:
        part_upper = part.strip().upper()
        all_parts_data.append({
            'customer_part_id': part,
            'in_quote_data': part_upper in quote_parts,
            'in_landmark': part_upper in landmark,
            'has_pd_match': part_upper in pd_cache,
        })

    total_rev_new = sum(r['mapped_med_rev'] for r in new_deals_data)
    total_rev_pd = sum(r['value'] for r in pd_info_data)
    won_deals = [r for r in pd_info_data if r['status'].lower() == 'won']
    open_deals = [r for r in pd_info_data if r['status'].lower() == 'open']

    status_counts = {}
    for r in pd_info_data:
        s = r['status'] or 'unknown'
        status_counts[s] = status_counts.get(s, 0) + 1

    platform_counts = {}
    platform_rev = {}
    for r in pd_info_data:
        p = r['platform_company'] or 'Unknown'
        platform_counts[p] = platform_counts.get(p, 0) + 1
        platform_rev[p] = platform_rev.get(p, 0) + r['value']

    label_counts = {}
    for r in pd_info_data:
        lb = r['label'] or 'Unknown'
        label_counts[lb] = label_counts.get(lb, 0) + 1

    stage_counts = {}
    for r in pd_info_data:
        st = r['stage_id'] or 'Unknown'
        stage_counts[st] = stage_counts.get(st, 0) + 1

    industry_counts = {}
    for r in pd_info_data:
        ind = r['industry'] or 'Unknown'
        if ind:
            industry_counts[ind] = industry_counts.get(ind, 0) + 1

    deal_type_counts = {}
    for r in pd_info_data:
        dt = r['deal_type'] or 'Unknown'
        if dt:
            deal_type_counts[dt] = deal_type_counts.get(dt, 0) + 1

    new_status_counts = {}
    for r in new_deals_data:
        s = r['mapped_status'] or 'Unknown'
        new_status_counts[s] = new_status_counts.get(s, 0) + 1

    calc_label_counts = {}
    for r in new_deals_data:
        cl = r['calc_label'] or 'Unknown'
        calc_label_counts[cl] = calc_label_counts.get(cl, 0) + 1

    top_customers_new = {}
    for r in new_deals_data:
        nm = r['name'] or 'Unknown'
        top_customers_new[nm] = top_customers_new.get(nm, 0) + r['mapped_med_rev']
    top_customers_new = sorted(top_customers_new.items(), key=lambda x: x[1], reverse=True)[:15]

    top_customers_pd = {}
    for r in pd_info_data:
        nm = r['org_name'] or 'Unknown'
        top_customers_pd[nm] = top_customers_pd.get(nm, 0) + r['value']
    top_customers_pd = sorted(top_customers_pd.items(), key=lambda x: x[1], reverse=True)[:15]

    json_result = {
        'output_file': OUTPUT_FILE,
        'natman_bookings_file': natman_path,
        'pdsync_file': pdsync_path,
        'elapsed_seconds': round(elapsed, 1),
        'summary': {
            'total_unique_parts': len(all_parts_sorted),
            'new_deals_count': len(new_deals_data),
            'pd_info_count': len(pd_info_data),
            'total_new_deals_revenue': round(total_rev_new, 2),
            'total_pd_pipeline_value': round(total_rev_pd, 2),
            'won_deals_count': len(won_deals),
            'won_deals_value': round(sum(r['value'] for r in won_deals), 2),
            'open_deals_count': len(open_deals),
            'open_deals_value': round(sum(r['value'] for r in open_deals), 2),
            'avg_deal_value_pd': round(total_rev_pd / len(pd_info_data), 2) if pd_info_data else 0,
            'avg_deal_value_new': round(total_rev_new / len(new_deals_data), 2) if new_deals_data else 0,
            'landmark_parts_count': len(landmark),
            'pd_cache_entries': len(pd_cache),
        },
        'analytics': {
            'pd_status_distribution': status_counts,
            'platform_distribution': platform_counts,
            'platform_revenue': {k: round(v, 2) for k, v in platform_rev.items()},
            'label_distribution': label_counts,
            'stage_distribution': stage_counts,
            'industry_distribution': industry_counts,
            'deal_type_distribution': deal_type_counts,
            'new_deals_status_distribution': new_status_counts,
            'calc_label_distribution': calc_label_counts,
            'top_customers_new': [{'name': n, 'revenue': round(r, 2)} for n, r in top_customers_new],
            'top_customers_pd': [{'name': n, 'value': round(r, 2)} for n, r in top_customers_pd],
        },
        'sheets': {
            'all_unique_parts': all_parts_data,
            'new_deals': new_deals_data,
            'pd_info': pd_info_data,
        },
        'source_data': {
            'bookings_sheets': natman_sheets,
            'national_sheets': pdsync_sheets,
        }
    }

    if args.json_output:
        with open(args.json_output, 'w') as jf:
            json.dump(json_result, jf, default=str)
        print(f"  JSON summary: {args.json_output}")

    return OUTPUT_FILE


if __name__ == '__main__':
    main()
